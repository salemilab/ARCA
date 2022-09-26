#!/usr/bin/env python3

import sys
import os.path
import shutil
import glob
from datetime import datetime
import sqlite3 as sql
from openpyxl import load_workbook

DBNAME = "arca.db"

DBCREATE = ["""CREATE TABLE Countries (
  idx integer PRIMARY KEY AUTOINCREMENT,
  subregion int,
  country text, code char(1) default 'C');""",
            """CREATE TABLE CanonicalNames (
  name text,
  canonical text);""",
            """CREATE TABLE Viruses (
  idx int PRIMARY KEY,
  virus text);""",
            """CREATE TABLE Cases (
  country int,
  virus int,
  year int,
  week int,
  cases int,
  totalcases int);""",
            """CREATE INDEX Cases_country ON Cases(country);""",
            """CREATE INDEX Cases_virus ON Cases(virus);"""]

def opendb():
    return sql.connect(DBNAME)

def get_excel_sheet(filename):
    wb = load_workbook(filename = filename)
    sheetname = wb.sheetnames[0]
    return wb[sheetname]

def convert_to_tab_delimited(filename):
    sh = get_excel_sheet(filename)

    for row in sh.rows:
        # Version 1:
        #for cell in row:
        #    sys.stdout.write(str(cell.value) + "\t")

        # Version 2:
        #sys.stdout.write(str(row[0].value))
        #for cell in row[1:]:
        #    sys.stdout.write("\t" + str(cell.value))
        #sys.stdout.write("\n")

        # Version 3:
        sys.stdout.write("\t".join([ str(c.value) for c in row ]) + "\n")

def store_countries(filename):
    sh = get_excel_sheet(filename)
    db = opendb()
    data = False
    savedid = 0
    success = False

    try:
        for row in sh.rows:
            if data:
                regid = row[0].value
                country = row[1].value
                if regid is None:
                    regid = savedid
                else:
                    savedid = regid
                db.execute("INSERT INTO Countries (subregion, country) VALUES (?, ?);", (regid, country))

            else:
                data = True
        success = True

    finally:
        if success:
            db.commit()
        db.close()

def get_country_id(db, country, regid, create=False):
    row = db.execute("SELECT canonical from CanonicalNames WHERE name=?", (country,)).fetchone()
    if row:
        sys.stderr.write("Replacing {} with {}.\n".format(country, row[0]))
        country = row[0]
    
    idx = db.execute("SELECT idx FROM Countries WHERE country=?", (country,)).fetchone()
    if idx:
        return idx[0]

    if create:
        try:
            db.execute("INSERT INTO Countries (subregion, country) VALUES (?, ?);", (regid, country))
            db.commit()
            idx = db.execute("SELECT idx FROM Countries WHERE country=?", (country,)).fetchone()
            return idx[0]
        except:
            return None
    else:
        sys.stderr.write("Unknown country: {}\n".format(country))
        
def get_totalcases(db, virus, country, year, week):
    row = db.execute("SELECT totalcases FROM Cases WHERE virus=? AND country=? AND year=? AND week<? ORDER BY week DESC LIMIT 1;", (virus, country, year, week)).fetchone()
    if row:
        #sys.stderr.write("{}, {}: {}\n".format(country, year, row[0]))
        return row[0]
    else:
        #sys.stderr.write("{}, {}: empty\n".format(country, year))
        return 0

def get_virus_id(db, name):
    row = db.execute("SELECT idx FROM Viruses WHERE virus=?;", (name,)).fetchone()
    if row:
        return row[0]
    else:
        sys.stderr.write("Unknown virus: {}\n".format(name))
        sys.exit(1)

def get_year_week(filename):
    name = os.path.split(filename)[1]
    year = name[:4]
    week = name[9:11]
    return year, week

def parse_one_file(db, virus, filename):
    year, week = get_year_week(filename)
    sh = get_excel_sheet(filename)
    data = False
    success = False
    rn = 0
    sumcases = 0
    
    try:
        for row in sh.rows:
            if data:
                rn += 1
                regid = row[0].value
                country = row[1].value
                countryid = get_country_id(db, country, regid)
                cases = row[6].value

                if cases is None:
                    continue
                totalcases = int(cases)
                
#                if cases is None:
#                    totalcases = 0
#                else:
#                    totalcases = int(cases)

                prev_totalcases = get_totalcases(db, virus, countryid, year, week)

                # If cumulative totals go down (why??), just ignore them to avoid
                # storing negative counts in the db.
                if totalcases < prev_totalcases:
                    continue
                
                newcases = totalcases - prev_totalcases
                sumcases += newcases
                db.execute("INSERT INTO Cases (country, virus, year, week, cases, totalcases) VALUES (?, ?, ?, ?, ?, ?);",
                           (countryid, virus, year, week, newcases, totalcases))
            else:
                data = True
        success = True

    finally:
        if success:
            sys.stderr.write("Parsing {}: success, {} rows read, {} cases\n".format(filename, rn, sumcases))
            db.commit()
        else:
            sys.stderr.write("Parsing {}: failed.\n".format(filename))

def parse_monthly_file(db, virus, filename):
    sh = get_excel_sheet(filename)
    rn = 0
    success = False
    try:
        for row in sh.rows:
            if rn == 0:
                years = row
            elif rn == 1:
                weeks = row
            elif rn > 2:
                regid = row[0].value
                country = row[1].value
                countryid = get_country_id(db, country, regid)
                totalcases = 0
                
                for i in range(2, 55):
                    year = years[i].value
                    if year is None:
                        totalcases = 0
                    else:
                        week = weeks[i].value
                        newcases = row[i].value
                        if newcases is None:
                            continue
                        totalcases += newcases
                        db.execute("INSERT INTO Cases (country, virus, year, week, cases, totalcases) VALUES (?, ?, ?, ?, ?, ?);",
                                   (countryid, virus, year, week, newcases, totalcases))

                #sys.stdout.write("{}\n{}\n".format([c.value for c in years], [c.value for c in weeks]))
                #sys.stdout.write("{}: {}\n".format(len(row), [c.value for c in row]))
                #for i in range(2,55):
                #    sys.stdout.write("{} {} {}\n".format(years[i].value, weeks[i].value, row[i].value))
            rn += 1
        success = True
    finally:
        if success:
            sys.stderr.write("Parsing {}: success, {} rows read\n".format(filename, rn-2))
            db.commit()
        else:
            sys.stderr.write("Parsing {}: failed.\n".format(filename))
        
def main_weekly(virus, filenames):
    db = opendb()
    try:
        virus_idx = get_virus_id(db, virus)
        for f in filenames:
            parse_one_file(db, virus_idx, f)
    finally:
        db.close()

def main_monthly(virus, filenames):
    db = opendb()
    try:
        virus_idx = get_virus_id(db, virus)
        for f in filenames:
            parse_monthly_file(db, virus_idx, f)
    finally:
        db.close()

def save_db():
    backup = DBNAME + "." + datetime.utcnow().strftime('%Y%m%d-%H:%M:%S')
    shutil.copyfile(DBNAME, backup)
        
# Create a new DB

def main_create():
    if os.path.isfile(DBNAME):
        sys.stderr.write("Database {} already exists!\n".format(DBNAME))
        return

    db = opendb()
    for sql in DBCREATE:
        db.execute(sql)
    db.close()
    sys.stderr.write("Database {} created.\n".format(DBNAME))

# Revert DB

def main_revert():
    patt = DBNAME + ".*-*:*:*"
    backups = glob.glob(patt)
    backups.sort(reverse=True)
    backup = backups[0]
    ans = input("Reverting to db version: {} - proceed? (y/N)\n".format(backup))
    if ans in "yY":
        shutil.copyfile(backup, DBNAME)

# Print unique country names

def main_countries(filenames):
    names = []
    data = False
    
    for filename in filenames:
        sh = get_excel_sheet(filename)
        for row in sh.rows:
            if data:
                country = row[1].value
                if country and (country not in names):
                    names.append(country)
            else:
                data = True
    #print(names)
    names.sort()
    for n in names:
        sys.stdout.write(n + "\n")

def usage():
    sys.stdout.write("""parser.py - Parser for ARCA database files.

Usage: parser.py command [options] args...

Where command is one of: create, load, revert, countries.

Options:

  -d DB | Use specified database instead of default one ({}).
  -m    | Monthly mode (for load).

Commands:

  create    - Create new database.

  load      - Load the specified files in the database. First argument should be 
              virus name (capitalized).

  backup    - Create a backup of the current database.

  revert    - Restore previous version of database (before the last load command).

  countries - Write list of countries appearing in the input files.

""".format(DBNAME))
        
def main(args):
    global DBNAME
    monthly_mode = False

    if len(args) == 0 or "-h" in args or "--help" in args:
        return usage()

    cmd = args[0]
    trueargs = []
    prev = ""
    for a in args[1:]:
        if prev == "-d":
            DBNAME = a
            prev = ""
        elif a == "-d":
            prev = a
        elif a == "-m":
            monthly_mode = True
        else:
            trueargs.append(a)
            
    if cmd == "load":
        if len(trueargs) < 2:
            return usage()
        virus = trueargs[0]
        filenames = trueargs[1:]
        save_db()
        if monthly_mode:
            main_monthly(virus, filenames)
        else:
            main_weekly(virus, filenames)

    elif cmd == "backup":
        save_db()
            
    elif cmd == "create":
        main_create()

    elif cmd == "revert":
        main_revert()

    elif cmd == "countries":
        main_countries(trueargs)
        
    else:
        return usage()

        
if __name__ == "__main__":
    main(sys.argv[1:])



    
#    virus = sys.argv[1]
#    filenames = sys.argv[2:]
#    main(virus, filenames)

    #main_monthly(virus, filenames[0])

    #convert_to_tab_delimited(sys.argv[1])
    #store_countries(sys.argv[1])
    #db = opendb()
    #idx = get_country_id(db, "Canada", 1)
    #print(idx)
    #idx = get_country_id(db, "UK", 99)
    #print(idx)


#parser.py load Zika f1 f2 f3...
#parser.py revert
#parser.py create
